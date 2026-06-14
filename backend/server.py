from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
import io
import random
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import google.generativeai as genai
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'change-me')
JWT_ALG = 'HS256'
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
genai.configure(api_key=GEMINI_API_KEY)

app = FastAPI(title="NextHeir API")
api_router = APIRouter(prefix="/api")


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False


def create_token(user_id: str) -> str:
    payload = {
        'sub': user_id,
        'iat': now_utc(),
        'exp': now_utc() + timedelta(days=30),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=401, detail='Missing or invalid token')
    token = authorization.split(' ', 1)[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        user_id = payload.get('sub')
    except Exception:
        raise HTTPException(status_code=401, detail='Invalid token')
    user = await db.users.find_one({'id': user_id}, {'_id': 0, 'password': 0})
    if not user:
        raise HTTPException(status_code=401, detail='User not found')
    return user


# ============ Models ============
class RegisterEmail(BaseModel):
    full_name: str
    email: EmailStr
    password: str


class LoginEmail(BaseModel):
    email: EmailStr
    password: str


class MobileRequest(BaseModel):
    full_name: Optional[str] = None
    mobile: str


class OTPVerify(BaseModel):
    mobile: str
    otp: str
    full_name: Optional[str] = None


class AssetIn(BaseModel):
    name: str
    category: str  # property | business | investment | precious_metal | other
    value: float
    description: Optional[str] = ''


class FamilyIn(BaseModel):
    name: str
    relationship: str
    age: Optional[int] = None
    financial_needs: Optional[str] = ''  # low | medium | high
    notes: Optional[str] = ''


class ScenarioIn(BaseModel):
    name: str
    description: Optional[str] = ''
    # allocations[asset_id][member_id] = percentage (0-100)
    allocations: Dict[str, Dict[str, float]] = {}


class ChatMessageIn(BaseModel):
    session_id: Optional[str] = None
    message: str
    scenario_ids: Optional[List[str]] = None


# ============ Auth Endpoints ============
@api_router.post('/auth/register')
async def register(data: RegisterEmail):
    existing = await db.users.find_one({'email': data.email})
    if existing:
        raise HTTPException(status_code=400, detail='Email already registered')
    user_id = str(uuid.uuid4())
    user_doc = {
        'id': user_id,
        'full_name': data.full_name,
        'email': data.email,
        'mobile': None,
        'password': hash_password(data.password),
        'auth_method': 'email',
        'currency': 'INR',
        'created_at': now_utc().isoformat(),
    }
    await db.users.insert_one(user_doc)
    token = create_token(user_id)
    return {
        'token': token,
        'user': {
            'id': user_id,
            'full_name': data.full_name,
            'email': data.email,
            'mobile': None,
            'currency': 'INR',
        },
    }


@api_router.post('/auth/login')
async def login(data: LoginEmail):
    user = await db.users.find_one({'email': data.email})
    if not user or not verify_password(data.password, user.get('password', '')):
        raise HTTPException(status_code=401, detail='Invalid credentials')
    token = create_token(user['id'])
    return {
        'token': token,
        'user': {
            'id': user['id'],
            'full_name': user.get('full_name'),
            'email': user.get('email'),
            'mobile': user.get('mobile'),
            'currency': user.get('currency', 'INR'),
            'role': user.get('role', 'user'),
        },
    }


@api_router.post('/auth/mobile/request-otp')
async def request_otp(data: MobileRequest):
    # Mocked OTP - always 123456 for demo, but return random for realism
    otp = '123456'
    await db.otps.update_one(
        {'mobile': data.mobile},
        {'$set': {'mobile': data.mobile, 'otp': otp, 'created_at': now_utc().isoformat()}},
        upsert=True,
    )
    return {'success': True, 'mobile': data.mobile, 'otp_mock': otp, 'message': 'OTP sent (mocked). Use 123456.'}


@api_router.post('/auth/mobile/verify-otp')
async def verify_otp(data: OTPVerify):
    rec = await db.otps.find_one({'mobile': data.mobile})
    if not rec or rec.get('otp') != data.otp:
        raise HTTPException(status_code=400, detail='Invalid OTP')
    user = await db.users.find_one({'mobile': data.mobile})
    if not user:
        user_id = str(uuid.uuid4())
        user = {
            'id': user_id,
            'full_name': data.full_name or 'User',
            'email': None,
            'mobile': data.mobile,
            'password': None,
            'auth_method': 'mobile',
            'currency': 'INR',
            'created_at': now_utc().isoformat(),
        }
        await db.users.insert_one(user)
    await db.otps.delete_one({'mobile': data.mobile})
    token = create_token(user['id'])
    return {
        'token': token,
        'user': {
            'id': user['id'],
            'full_name': user.get('full_name'),
            'email': user.get('email'),
            'mobile': user.get('mobile'),
            'currency': user.get('currency', 'INR'),
        },
    }


class CurrencyUpdate(BaseModel):
    currency: str


@api_router.patch('/auth/me/currency')
async def update_currency(data: CurrencyUpdate, user=Depends(get_current_user)):
    allowed = {'INR', 'USD', 'EUR', 'GBP', 'AED', 'SGD', 'JPY', 'AUD', 'CAD', 'CHF', 'CNY', 'HKD'}
    if data.currency not in allowed:
        raise HTTPException(status_code=400, detail='Unsupported currency')
    await db.users.update_one({'id': user['id']}, {'$set': {'currency': data.currency}})
    return {'success': True, 'currency': data.currency}


@api_router.get('/auth/me')
async def me(user=Depends(get_current_user)):
    return user


# ============ Assets ============
@api_router.get('/assets')
async def list_assets(user=Depends(get_current_user)):
    items = await db.assets.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    return items


@api_router.post('/assets')
async def create_asset(data: AssetIn, user=Depends(get_current_user)):
    item = {
        'id': str(uuid.uuid4()),
        'user_id': user['id'],
        'name': data.name,
        'category': data.category,
        'value': data.value,
        'description': data.description,
        'created_at': now_utc().isoformat(),
    }
    await db.assets.insert_one(item)
    item.pop('_id', None)
    return item


@api_router.put('/assets/{asset_id}')
async def update_asset(asset_id: str, data: AssetIn, user=Depends(get_current_user)):
    res = await db.assets.update_one(
        {'id': asset_id, 'user_id': user['id']},
        {'$set': data.dict()},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail='Asset not found')
    item = await db.assets.find_one({'id': asset_id}, {'_id': 0})
    return item


@api_router.delete('/assets/{asset_id}')
async def delete_asset(asset_id: str, user=Depends(get_current_user)):
    await db.assets.delete_one({'id': asset_id, 'user_id': user['id']})
    return {'success': True}


# ============ Family Members ============
@api_router.get('/family')
async def list_family(user=Depends(get_current_user)):
    items = await db.family.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    return items


@api_router.post('/family')
async def create_family(data: FamilyIn, user=Depends(get_current_user)):
    item = {
        'id': str(uuid.uuid4()),
        'user_id': user['id'],
        **data.dict(),
        'created_at': now_utc().isoformat(),
    }
    await db.family.insert_one(item)
    item.pop('_id', None)
    return item


@api_router.put('/family/{member_id}')
async def update_family(member_id: str, data: FamilyIn, user=Depends(get_current_user)):
    res = await db.family.update_one(
        {'id': member_id, 'user_id': user['id']},
        {'$set': data.dict()},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail='Member not found')
    return await db.family.find_one({'id': member_id}, {'_id': 0})


@api_router.delete('/family/{member_id}')
async def delete_family(member_id: str, user=Depends(get_current_user)):
    await db.family.delete_one({'id': member_id, 'user_id': user['id']})
    return {'success': True}


# ============ Scenarios ============
@api_router.get('/scenarios')
async def list_scenarios(user=Depends(get_current_user)):
    items = await db.scenarios.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    return items


@api_router.get('/scenarios/{scenario_id}')
async def get_scenario(scenario_id: str, user=Depends(get_current_user)):
    item = await db.scenarios.find_one({'id': scenario_id, 'user_id': user['id']}, {'_id': 0})
    if not item:
        raise HTTPException(status_code=404, detail='Scenario not found')
    return item


@api_router.post('/scenarios')
async def create_scenario(data: ScenarioIn, user=Depends(get_current_user)):
    item = {
        'id': str(uuid.uuid4()),
        'user_id': user['id'],
        'name': data.name,
        'description': data.description,
        'allocations': data.allocations,
        'analysis': None,
        'share_token': None,
        'created_at': now_utc().isoformat(),
        'updated_at': now_utc().isoformat(),
    }
    await db.scenarios.insert_one(item)
    item.pop('_id', None)
    return item


@api_router.put('/scenarios/{scenario_id}')
async def update_scenario(scenario_id: str, data: ScenarioIn, user=Depends(get_current_user)):
    res = await db.scenarios.update_one(
        {'id': scenario_id, 'user_id': user['id']},
        {'$set': {**data.dict(), 'updated_at': now_utc().isoformat(), 'analysis': None}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail='Scenario not found')
    return await db.scenarios.find_one({'id': scenario_id}, {'_id': 0})


@api_router.delete('/scenarios/{scenario_id}')
async def delete_scenario(scenario_id: str, user=Depends(get_current_user)):
    await db.scenarios.delete_one({'id': scenario_id, 'user_id': user['id']})
    return {'success': True}


# ============ AI: Fairness Analysis ============
@api_router.post('/scenarios/{scenario_id}/analyze')
async def analyze_scenario(scenario_id: str, user=Depends(get_current_user)):
    scenario = await db.scenarios.find_one({'id': scenario_id, 'user_id': user['id']}, {'_id': 0})
    if not scenario:
        raise HTTPException(status_code=404, detail='Scenario not found')

    assets = await db.assets.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    family = await db.family.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)

    asset_map = {a['id']: a for a in assets}
    member_map = {m['id']: m for m in family}

    # Compute per-member totals
    member_totals: Dict[str, float] = {m['id']: 0.0 for m in family}
    total_value = sum(a.get('value', 0) for a in assets)
    for asset_id, alloc in (scenario.get('allocations') or {}).items():
        asset = asset_map.get(asset_id)
        if not asset:
            continue
        for member_id, pct in alloc.items():
            if member_id in member_totals:
                member_totals[member_id] += (asset.get('value', 0) * float(pct) / 100.0)

    summary_lines = []
    for mid, total in member_totals.items():
        m = member_map.get(mid, {})
        share_pct = (total / total_value * 100) if total_value else 0
        summary_lines.append(
            f"- {m.get('name', 'Unknown')} ({m.get('relationship', '')}, age {m.get('age', 'N/A')}, "
            f"financial needs: {m.get('financial_needs', 'unknown')}): receives ₹{total:,.0f} ({share_pct:.1f}%)"
        )

    prompt = f"""You are NextHeir AI, an inheritance planning advisor. Analyze this distribution scenario and return a strict JSON object.

SCENARIO: {scenario.get('name')}
Description: {scenario.get('description', '')}
Total Estate Value: ₹{total_value:,.0f}
Number of Family Members: {len(family)}
Number of Assets: {len(assets)}

DISTRIBUTION SUMMARY:
{chr(10).join(summary_lines) if summary_lines else 'No allocations defined.'}

Return ONLY a valid JSON object with these exact keys (no markdown, no backticks):
{{
  "fairness_score": <integer 0-100>,
  "fairness_label": "<High|Medium|Low>",
  "summary": "<2-3 sentence overall assessment>",
  "strengths": ["<strength 1>", "<strength 2>"],
  "risks": [
    {{"level": "<high|medium|low>", "title": "<short title>", "detail": "<one sentence>"}}
  ],
  "recommendations": ["<actionable suggestion 1>", "<actionable suggestion 2>", "<actionable suggestion 3>"]
}}

Consider equality, need-based fairness, conflict risk between siblings, financial vulnerability, and long-term family harmony."""

    try:
        model = genai.GenerativeModel(
    model_name="gemini-3-flash-preview",
    system_instruction="You are NextHeir AI, an expert inheritance and wealth distribution advisor. Always respond with valid JSON only when requested."
)

response = model.generate_content(prompt)
text = response.text.strip()
        # Strip code fences if any
        if text.startswith('```'):
            text = text.split('```', 2)[1]
            if text.startswith('json'):
                text = text[4:]
            text = text.strip().rstrip('```').strip()
        analysis = json.loads(text)
    except Exception as e:
        logging.exception("AI analysis failed: %s", e)
        analysis = {
            "fairness_score": 50,
            "fairness_label": "Medium",
            "summary": "Unable to perform full AI analysis at this time. Please review the distribution manually.",
            "strengths": [],
            "risks": [{"level": "medium", "title": "AI analysis unavailable", "detail": "Try regenerating the analysis."}],
            "recommendations": ["Add more details to your assets and family members.", "Try analyzing again."],
        }

    analysis['totals_by_member'] = {
        mid: {
            'name': member_map.get(mid, {}).get('name', 'Unknown'),
            'amount': member_totals[mid],
            'percentage': (member_totals[mid] / total_value * 100) if total_value else 0,
        }
        for mid in member_totals
    }
    analysis['total_estate_value'] = total_value

    await db.scenarios.update_one(
        {'id': scenario_id, 'user_id': user['id']},
        {'$set': {'analysis': analysis, 'updated_at': now_utc().isoformat()}},
    )
    return analysis


# ============ AI Chat ============
@api_router.post('/chat')
async def chat_message(data: ChatMessageIn, user=Depends(get_current_user)):
    session_id = data.session_id or str(uuid.uuid4())

    # Build context from user's data
    assets = await db.assets.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    family = await db.family.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    total_value = sum(a.get('value', 0) for a in assets)
    asset_map = {a['id']: a for a in assets}
    member_map = {m['id']: m for m in family}

    context = f"""User's Estate Snapshot:
- Total estate value: ₹{total_value:,.0f}
- Assets ({len(assets)}): {', '.join([f"{a.get('name')} ({a.get('category')}, ₹{a.get('value', 0):,.0f})" for a in assets[:10]]) or 'none yet'}
- Family members ({len(family)}): {', '.join([f"{m.get('name')} ({m.get('relationship')}, needs: {m.get('financial_needs', 'unknown')})" for m in family[:10]]) or 'none yet'}
"""

    # If user has selected scenarios for comparison, include detailed allocations + analysis
    scenarios_block = ""
    if data.scenario_ids:
        selected = await db.scenarios.find(
            {'id': {'$in': data.scenario_ids}, 'user_id': user['id']},
            {'_id': 0},
        ).to_list(50)
        if selected:
            parts = []
            for s in selected:
                lines = [f"\n=== SCENARIO: \"{s.get('name')}\" ==="]
                if s.get('description'):
                    lines.append(f"Description: {s['description']}")
                # Per-member totals
                m_totals: Dict[str, float] = {m['id']: 0.0 for m in family}
                for aid, alloc in (s.get('allocations') or {}).items():
                    a = asset_map.get(aid)
                    if not a:
                        continue
                    for mid, pct in alloc.items():
                        if mid in m_totals:
                            m_totals[mid] += a.get('value', 0) * float(pct) / 100.0
                lines.append("Distribution:")
                for mid, total in m_totals.items():
                    if total <= 0:
                        continue
                    m = member_map.get(mid, {})
                    share = (total / total_value * 100) if total_value else 0
                    lines.append(f"  - {m.get('name', 'Unknown')} ({m.get('relationship', '')}): ₹{total:,.0f} ({share:.1f}%)")
                ana = s.get('analysis')
                if ana:
                    lines.append(f"AI Fairness Score: {ana.get('fairness_score', 'N/A')}/100 ({ana.get('fairness_label', '')})")
                    if ana.get('summary'):
                        lines.append(f"Prior AI Summary: {ana['summary']}")
                else:
                    lines.append("(Not yet analyzed)")
                parts.append('\n'.join(lines))
            scenarios_block = "\n\nSCENARIOS SELECTED FOR COMPARISON:\n" + '\n'.join(parts) + "\n\nWhen the user asks a comparison question, contrast these scenarios across fairness, conflict risk, financial need coverage, and family harmony. Recommend the strongest option with reasoning, then highlight one trade-off."

    system_msg = f"""You are NextHeir AI, a warm and insightful inheritance planning advisor for high-net-worth families.
You help users think through fairness, emotional dynamics, conflict prevention, and long-term family harmony.

CRITICAL FORMATTING RULE: Every single reply you send MUST begin with this exact disclaimer line as the first line, in italics, followed by a blank line:
"_Disclaimer: This is not financial advice. Please consult your Chartered Accountant, financial advisor, or wealth manager before making any inheritance distribution decisions._"

After the disclaimer, answer the user's question. Keep replies concise (under 220 words), use bullets when helpful, and always end with a thoughtful follow-up question.
Never provide legal or tax advice — always remind users to consult their CA or lawyer for final decisions.

{context}{scenarios_block}"""

    try:
        model = genai.GenerativeModel(
    model_name="gemini-3-flash-preview",
    system_instruction=system_msg
)

response = model.generate_content(data.message)
response_text = response.text
    except Exception as e:
        logging.exception("AI chat failed: %s", e)
        response_text = "I'm having trouble connecting right now. Please try again in a moment."

    ts = now_utc().isoformat()
    await db.chat_messages.insert_many([
        {'id': str(uuid.uuid4()), 'user_id': user['id'], 'session_id': session_id, 'role': 'user', 'content': data.message, 'created_at': ts},
        {'id': str(uuid.uuid4()), 'user_id': user['id'], 'session_id': session_id, 'role': 'assistant', 'content': response_text, 'created_at': now_utc().isoformat()},
    ])

    return {'session_id': session_id, 'reply': response_text}


@api_router.get('/chat/history')
async def chat_history(session_id: Optional[str] = None, user=Depends(get_current_user)):
    q = {'user_id': user['id']}
    if session_id:
        q['session_id'] = session_id
    msgs = await db.chat_messages.find(q, {'_id': 0}).sort('created_at', 1).to_list(500)
    return msgs


@api_router.delete('/chat/history')
async def clear_chat(session_id: Optional[str] = None, user=Depends(get_current_user)):
    q = {'user_id': user['id']}
    if session_id:
        q['session_id'] = session_id
    await db.chat_messages.delete_many(q)
    return {'success': True}


# ============ Share Scenario ============
@api_router.post('/scenarios/{scenario_id}/share')
async def share_scenario(scenario_id: str, user=Depends(get_current_user)):
    scenario = await db.scenarios.find_one({'id': scenario_id, 'user_id': user['id']})
    if not scenario:
        raise HTTPException(status_code=404, detail='Scenario not found')
    token = scenario.get('share_token') or str(uuid.uuid4()).replace('-', '')[:16]
    await db.scenarios.update_one({'id': scenario_id}, {'$set': {'share_token': token}})
    return {'share_token': token, 'share_url': f'/api/shared/{token}'}


@api_router.get('/shared/{token}')
async def get_shared(token: str):
    scenario = await db.scenarios.find_one({'share_token': token}, {'_id': 0})
    if not scenario:
        raise HTTPException(status_code=404, detail='Shared scenario not found')
    user = await db.users.find_one({'id': scenario['user_id']}, {'_id': 0, 'password': 0})
    assets = await db.assets.find({'user_id': scenario['user_id']}, {'_id': 0}).to_list(1000)
    family = await db.family.find({'user_id': scenario['user_id']}, {'_id': 0}).to_list(1000)
    return {
        'scenario': scenario,
        'owner_name': user.get('full_name') if user else 'Anonymous',
        'assets': assets,
        'family': family,
    }


# ============ PDF Export ============
@api_router.get('/scenarios/{scenario_id}/pdf')
async def scenario_pdf(scenario_id: str, user=Depends(get_current_user)):
    scenario = await db.scenarios.find_one({'id': scenario_id, 'user_id': user['id']}, {'_id': 0})
    if not scenario:
        raise HTTPException(status_code=404, detail='Scenario not found')
    assets = await db.assets.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    family = await db.family.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    asset_map = {a['id']: a for a in assets}
    member_map = {m['id']: m for m in family}

    code = user.get('currency', 'INR')
    symbols = {'INR': 'Rs.', 'USD': '$', 'EUR': 'EUR ', 'GBP': 'GBP ', 'AED': 'AED ', 'SGD': 'S$', 'JPY': 'JPY ', 'AUD': 'A$', 'CAD': 'C$', 'CHF': 'CHF ', 'CNY': 'CNY ', 'HKD': 'HK$'}
    sym = symbols.get(code, 'Rs.')

    def fmt(n):
        return f"{sym}{(n or 0):,.0f}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=40)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], textColor=HexColor('#0A0A0E'), fontSize=24, spaceAfter=4)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], textColor=HexColor('#8C92AC'), fontSize=11, spaceAfter=20)
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], textColor=HexColor('#0A0A0E'), fontSize=15, spaceBefore=12, spaceAfter=8)
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=14)

    story = []
    story.append(Paragraph("NextHeir Distribution Report", title_style))
    story.append(Paragraph(f"Scenario: {scenario.get('name')}", sub_style))
    if scenario.get('description'):
        story.append(Paragraph(scenario['description'], body_style))
        story.append(Spacer(1, 12))

    total_value = sum(a.get('value', 0) for a in assets)
    story.append(Paragraph(f"<b>Total Estate Value:</b> {fmt(total_value)}", body_style))
    story.append(Paragraph(f"<b>Generated:</b> {now_utc().strftime('%d %b %Y, %H:%M UTC')}", body_style))
    story.append(Spacer(1, 12))

    # Allocations table
    story.append(Paragraph("Allocations", h2_style))
    headers = ['Asset', 'Value'] + [m['name'] for m in family]
    data_rows = [headers]
    for a in assets:
        row = [a['name'], fmt(a.get('value', 0))]
        alloc_for_asset = (scenario.get('allocations') or {}).get(a['id'], {})
        for m in family:
            pct = alloc_for_asset.get(m['id'], 0)
            row.append(f"{pct:.0f}%" if pct else '—')
        data_rows.append(row)
    if len(data_rows) > 1:
        table = Table(data_rows, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#0A0A0E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#D4AF37')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FAFAFA'), HexColor('#FFFFFF')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)

    # Per member totals
    story.append(Paragraph("Distribution Summary", h2_style))
    member_totals = []
    for m in family:
        total = 0.0
        for asset_id, alloc in (scenario.get('allocations') or {}).items():
            asset = asset_map.get(asset_id)
            if not asset:
                continue
            total += asset.get('value', 0) * float(alloc.get(m['id'], 0)) / 100.0
        share = (total / total_value * 100) if total_value else 0
        member_totals.append([m['name'], m.get('relationship', ''), fmt(total), f"{share:.1f}%"])
    if member_totals:
        t2 = Table([['Member', 'Relationship', 'Inheritance', 'Share %']] + member_totals, repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#0A0A0E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#D4AF37')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FAFAFA'), HexColor('#FFFFFF')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(t2)

    # AI Analysis
    analysis = scenario.get('analysis')
    if analysis:
        story.append(Paragraph("AI Fairness Analysis", h2_style))
        story.append(Paragraph(f"<b>Fairness Score:</b> {analysis.get('fairness_score', 0)} / 100 ({analysis.get('fairness_label', 'N/A')})", body_style))
        story.append(Spacer(1, 4))
        story.append(Paragraph(analysis.get('summary', ''), body_style))
        story.append(Spacer(1, 8))
        if analysis.get('strengths'):
            story.append(Paragraph("<b>Strengths</b>", body_style))
            for s in analysis['strengths']:
                story.append(Paragraph(f"• {s}", body_style))
        if analysis.get('risks'):
            story.append(Spacer(1, 6))
            story.append(Paragraph("<b>Risks</b>", body_style))
            for r in analysis['risks']:
                story.append(Paragraph(f"• [{r.get('level', '').upper()}] {r.get('title', '')}: {r.get('detail', '')}", body_style))
        if analysis.get('recommendations'):
            story.append(Spacer(1, 6))
            story.append(Paragraph("<b>Recommendations</b>", body_style))
            for r in analysis['recommendations']:
                story.append(Paragraph(f"• {r}", body_style))

    story.append(Spacer(1, 24))
    disclaimer = ParagraphStyle('Disclaimer', parent=styles['Normal'], fontSize=8, textColor=HexColor('#8C92AC'), italic=True)
    story.append(Paragraph(
        "Disclaimer: This report is AI-generated based on the inputs provided. It is indicative in nature and should not be considered financial or legal advice. Please consult your CA, lawyer, or wealth manager before making any final decisions.",
        disclaimer,
    ))

    doc.build(story)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="nextheir-{scenario.get("name", "scenario")}.pdf"'},
    )


# ============ Dashboard summary ============
@api_router.get('/dashboard')
async def dashboard(user=Depends(get_current_user)):
    assets = await db.assets.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    family = await db.family.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    scenarios = await db.scenarios.find({'user_id': user['id']}, {'_id': 0}).to_list(1000)
    total_value = sum(a.get('value', 0) for a in assets)
    by_category: Dict[str, float] = {}
    for a in assets:
        by_category[a['category']] = by_category.get(a['category'], 0) + a.get('value', 0)
    avg_score = None
    scored = [s.get('analysis', {}).get('fairness_score') for s in scenarios if s.get('analysis')]
    if scored:
        avg_score = sum(scored) / len(scored)
    return {
        'total_value': total_value,
        'asset_count': len(assets),
        'family_count': len(family),
        'scenario_count': len(scenarios),
        'by_category': by_category,
        'avg_fairness_score': avg_score,
    }


@api_router.get('/')
async def root():
    return {'message': 'NextHeir API', 'status': 'ok'}


# ============ Admin ============
ADMIN_EMAIL = 'admin@nextheir.com'
ADMIN_PASSWORD = 'Admin@123'


async def require_admin(user=Depends(get_current_user)):
    if user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail='Admin access required')
    return user


@api_router.get('/admin/users')
async def admin_list_users(admin=Depends(require_admin)):
    users = await db.users.find({}, {'_id': 0, 'password': 0}).sort('created_at', -1).to_list(5000)
    return users


@api_router.get('/admin/stats')
async def admin_stats(admin=Depends(require_admin)):
    total_users = await db.users.count_documents({})
    total_assets = await db.assets.count_documents({})
    total_family = await db.family.count_documents({})
    total_scenarios = await db.scenarios.count_documents({})
    total_chats = await db.chat_messages.count_documents({})
    by_method = {}
    async for u in db.users.find({}, {'auth_method': 1, '_id': 0}):
        by_method[u.get('auth_method', 'unknown')] = by_method.get(u.get('auth_method', 'unknown'), 0) + 1
    return {
        'total_users': total_users,
        'total_assets': total_assets,
        'total_family': total_family,
        'total_scenarios': total_scenarios,
        'total_chats': total_chats,
        'by_auth_method': by_method,
    }


@api_router.get('/admin/users/export')
async def admin_export_users(admin=Depends(require_admin)):
    users = await db.users.find({}, {'_id': 0, 'password': 0}).sort('created_at', -1).to_list(10000)
    wb = Workbook()
    ws = wb.active
    ws.title = 'NextHeir Users'

    headers = ['User Name', 'Email ID', 'Phone Number', 'Sign Up Date', 'Auth Method', 'Currency', 'Role']
    ws.append(headers)
    header_fill = PatternFill(start_color='0A0A0E', end_color='0A0A0E', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='D4AF37')
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for u in users:
        created = u.get('created_at', '')
        try:
            if created:
                # Format ISO timestamp -> "13 May 2026 06:43 UTC"
                dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                created = dt.strftime('%d %b %Y, %H:%M UTC')
        except Exception:
            pass
        ws.append([
            u.get('full_name', '') or '',
            u.get('email', '') or '',
            u.get('mobile', '') or '',
            created,
            u.get('auth_method', '') or '',
            u.get('currency', 'INR') or 'INR',
            u.get('role', 'user') or 'user',
        ])

    widths = [24, 30, 18, 24, 14, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'nextheir-users-{now_utc().strftime("%Y%m%d-%H%M")}.xlsx'
    return Response(
        content=buf.read(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.on_event("startup")
async def seed_admin():
    try:
        existing = await db.users.find_one({'email': ADMIN_EMAIL})
        if existing:
            # Ensure role is admin and password is current
            await db.users.update_one(
                {'email': ADMIN_EMAIL},
                {'$set': {'role': 'admin', 'password': hash_password(ADMIN_PASSWORD)}},
            )
        else:
            await db.users.insert_one({
                'id': str(uuid.uuid4()),
                'full_name': 'NextHeir Admin',
                'email': ADMIN_EMAIL,
                'mobile': None,
                'password': hash_password(ADMIN_PASSWORD),
                'auth_method': 'email',
                'role': 'admin',
                'currency': 'INR',
                'created_at': now_utc().isoformat(),
            })
        logging.info("Admin user ensured: %s", ADMIN_EMAIL)
    except Exception as e:
        logging.exception("Failed to seed admin: %s", e)


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
